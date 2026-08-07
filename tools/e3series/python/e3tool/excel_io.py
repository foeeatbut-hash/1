"""Чтение и запись .xlsx через openpyxl.

Установленный Microsoft Excel не нужен — ни для выгрузки, ни для загрузки.
Это единственное место, где программа знает про формат файла; экспорт и импорт
работают со списками словарей.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from typing import Any, Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import columns as cols
from .util import Row, Table, header_key, safe_str

HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
DATA_FONT = Font(size=11)
THIN = Side(style="thin", color="DCDCDC")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# --- оформление листа-отчёта ---------------------------------------------------
TITLE_FILL = PatternFill("solid", fgColor="DCE6F1")
TITLE_FONT = Font(bold=True, size=12, color="1F3864")
BLOCK_HEADER_FILL = PatternFill("solid", fgColor="C6D9F1")
BLOCK_HEADER_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill("solid", fgColor="EDEDED")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
NOTE_FONT = Font(italic=True, color="1D7A46")

#: Границы рисуются только для небольших листов: на десятках тысяч строк
#: они раздувают файл, не добавляя пользы.
BORDER_ROW_LIMIT = 5000

MIN_WIDTH = 10
MAX_WIDTH = 30


@dataclass
class SheetData:
    """Лист книги: заголовки, строки-словари и какие столбцы числовые."""

    name: str
    headers: list[str]
    rows: list[dict[str, Any]] = field(default_factory=list)
    numeric: set[str] = field(default_factory=set)


@dataclass
class Block:
    """Блок листа-отчёта: заголовок, шапка и строки списками.

    Отчёт — не таблица данных: в нём несколько разных таблиц одна под другой,
    поэтому строки здесь списками, а не словарями, и столбцы у каждого блока
    свои.
    """

    title: str = ""
    headers: list[str] = field(default_factory=list)
    rows: list[list[Any]] = field(default_factory=list)
    #: Индекс столбца со статусом: благополучное значение зелёное, прочее красное.
    status_column: int | None = None
    #: Какие значения статуса считать благополучными. Пусто — всё, что с «OK».
    ok_values: set[str] = field(default_factory=set)
    #: Сколько последних строк выделить как итоговые.
    total_rows: int = 0
    #: Что написать, если строк нет вовсе.
    empty_note: str = ""


@dataclass
class ReportSheet:
    """Лист книги, собранный из блоков. Читается человеком, не программой."""

    name: str
    blocks: list[Block] = field(default_factory=list)


# ------------------------------------------------------------------------------
#  Чтение
# ------------------------------------------------------------------------------
def read_tables(path: str, sheet_names: Iterable[str] | None = None) -> dict[str, Table]:
    """Читает книгу целиком: {имя листа -> Table}.

    Заголовки берутся из первой строки. Столбцы сопоставляются по имени, поэтому
    файл со старым набором столбцов читается без правок, а порядок не важен.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    wanted = None if sheet_names is None else {name.strip().lower() for name in sheet_names}
    tables: dict[str, Table] = {}
    try:
        for worksheet in workbook.worksheets:
            if wanted is not None and worksheet.title.strip().lower() not in wanted:
                continue
            table = _read_worksheet(worksheet)
            if table is not None:
                tables[worksheet.title] = table
    finally:
        workbook.close()
    return tables


def _read_worksheet(worksheet: Any) -> Table | None:
    iterator = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return None

    headers: list[str] = []
    positions: list[int] = []
    for index, value in enumerate(header_row):
        title = safe_str(value).strip()
        if title:
            headers.append(title)
            positions.append(index)
    if not headers:
        return None

    rows: list[Row] = []
    number = 1
    for raw in iterator:
        number += 1
        if raw is None:
            continue
        data: dict[str, Any] = {}
        empty = True
        for title, index in zip(headers, positions):
            value = raw[index] if index < len(raw) else None
            if value is not None and safe_str(value).strip() != "":
                empty = False
            data[header_key(title)] = value
        if empty:
            continue
        rows.append(Row(data, number))
    return Table(worksheet.title, headers, rows)


def find_table(tables: dict[str, Table], name: str) -> Table | None:
    """Находит лист по имени без учёта регистра."""
    target = name.strip().lower()
    for title, table in tables.items():
        if title.strip().lower() == target:
            return table
    return None


# ------------------------------------------------------------------------------
#  Запись
# ------------------------------------------------------------------------------
def write_workbook(path: str, sheets: Sequence[SheetData | ReportSheet]) -> str:
    """Пишет книгу и возвращает путь к сохранённому файлу."""
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    for data in sheets:
        worksheet = workbook.create_sheet(title=data.name[:31])
        if isinstance(data, ReportSheet):
            _write_report(worksheet, data)
        else:
            _write_sheet(worksheet, data)

    directory = os.path.dirname(os.path.abspath(path))
    if directory:
        os.makedirs(directory, exist_ok=True)
    workbook.save(path)
    return path


def _write_sheet(worksheet: Any, data: SheetData) -> None:
    worksheet.append(data.headers)

    numeric_columns = {
        index + 1 for index, header in enumerate(data.headers) if header in data.numeric
    }
    widths = [len(header) + 2 for header in data.headers]

    for row in data.rows:
        values = []
        for index, header in enumerate(data.headers):
            value = row.get(header)
            values.append(value)
            length = len(safe_str(value))
            if length + 2 > widths[index]:
                widths[index] = length + 2
        worksheet.append(values)

    row_count = len(data.rows)
    last_row = row_count + 1

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
    worksheet.row_dimensions[1].height = 24

    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(width, MIN_WIDTH), MAX_WIDTH
        )

    if numeric_columns and row_count:
        centre = Alignment(horizontal="center")
        for column in numeric_columns:
            letter = get_column_letter(column)
            for row_index in range(2, last_row + 1):
                cell = worksheet[f"{letter}{row_index}"]
                cell.number_format = "0.00"
                cell.alignment = centre

    if row_count and row_count <= BORDER_ROW_LIMIT:
        for row in worksheet.iter_rows(
            min_row=1, max_row=last_row, min_col=1, max_col=len(data.headers)
        ):
            for cell in row:
                cell.border = CELL_BORDER

    worksheet.freeze_panes = "A2"
    if row_count:
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(data.headers))}{last_row}"
        )


def _write_report(worksheet: Any, data: ReportSheet) -> None:
    """Пишет лист-отчёт: блоки один под другим, между ними пустая строка."""
    row = 1
    widths: dict[int, int] = {}

    def note_width(column: int, value: Any) -> None:
        length = len(safe_str(value)) + 2
        if length > widths.get(column, 0):
            widths[column] = length

    for block in data.blocks:
        if block.title:
            cell = worksheet.cell(row=row, column=1, value=block.title)
            cell.font = TITLE_FONT
            cell.fill = TITLE_FILL
            span = max(len(block.headers), 1)
            if span > 1:
                worksheet.merge_cells(
                    start_row=row, start_column=1, end_row=row, end_column=span
                )
            row += 1

        if not block.rows and block.empty_note:
            cell = worksheet.cell(row=row, column=1, value=block.empty_note)
            cell.font = NOTE_FONT
            row += 2
            continue

        for column, header in enumerate(block.headers, start=1):
            cell = worksheet.cell(row=row, column=column, value=header)
            cell.font = BLOCK_HEADER_FONT
            cell.fill = BLOCK_HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = CELL_BORDER
            note_width(column, header)
        row += 1

        first_total = len(block.rows) - block.total_rows
        for index, values in enumerate(block.rows):
            is_total = index >= first_total
            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(row=row, column=column, value=value)
                cell.border = CELL_BORDER
                note_width(column, value)
                if is_total:
                    cell.font = BLOCK_HEADER_FONT
                    cell.fill = TOTAL_FILL
                if block.status_column is not None and column == block.status_column + 1:
                    text = safe_str(value)
                    if text:
                        good = (
                            text in block.ok_values
                            if block.ok_values
                            else text.startswith("OK")
                        )
                        cell.fill = OK_FILL if good else BAD_FILL
            row += 1
        row += 1

    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = min(
            max(width, MIN_WIDTH), MAX_WIDTH
        )


# ------------------------------------------------------------------------------
#  Шаблон
# ------------------------------------------------------------------------------
def write_template(path: str) -> str:
    """Создаёт пустой шаблон со всеми столбцами и строками-примерами.

    Набор листов тот же, что у выгрузки: «Изделия», «Схема», «Подвал»,
    «Соединения», «Листы». Примеры показывают, что вид листа заполняется всегда —
    иначе лист с повторяющимся именем не определить.
    """
    example: dict[str, Any] = {header: "" for header in cols.DEVICE_HEADERS}
    example[cols.H_POZ] = "ВЕНТ-01-DI-001"
    example[cols.H_COMP] = "внешние_сигналы_DI"
    example["*TAG установки"] = "ВЕНТ-01"
    example["*Описание полное"] = "Пример строки: замените на свои данные"
    example[cols.H_SHEET] = "1"
    example[cols.H_VIEW] = "4"
    example[cols.H_X] = 100
    example[cols.H_Y] = 300

    def placement(zone: str, y: float) -> dict[str, Any]:
        row: dict[str, Any] = {header: "" for header in cols.PLACEMENT_HEADERS}
        row[cols.H_POZ] = "ВЕНТ-01-DI-001"
        # GID оставлен пустым намеренно: в шаблоне его взять неоткуда, он
        # появляется при выгрузке из проекта и там же становится главным ключом.
        row[cols.H_SYM_NR] = 1
        row[cols.H_SYM_DB] = "имя символа в базе E3"
        row[cols.H_SHEET] = "1"
        row[cols.H_VIEW] = "4"
        row[cols.H_ZONE] = zone
        row[cols.H_X] = 100
        row[cols.H_Y] = y
        return row

    sheet_example: dict[str, Any] = {header: "" for header in cols.SHEET_HEADERS}
    sheet_example[cols.H_SHEET] = "1"
    sheet_example[cols.H_VIEW] = "4"
    sheet_example[cols.H_VIEW_NAME] = cols.view_title("4")
    sheet_example[cols.H_FORMAT] = "A2_ГОСТ"

    text_example: dict[str, Any] = {header: "" for header in cols.TEXT_HEADERS}
    text_example[cols.H_SHEET] = "1"
    text_example[cols.H_VIEW] = "4"
    text_example[cols.H_X] = 40
    text_example[cols.H_Y] = 120
    text_example[cols.H_TEXT] = "Пример свободной надписи на листе"

    sheets: list[SheetData | ReportSheet] = [
        SheetData(
            name=cols.SHEET_DEVICES,
            headers=cols.DEVICE_HEADERS,
            rows=[example],
            numeric=set(cols.NUMERIC_HEADERS),
        ),
        SheetData(
            name=cols.SHEET_VIEW4,
            headers=cols.VIEW_DEVICE_HEADERS,
            rows=[example],
            numeric=set(cols.NUMERIC_HEADERS),
        ),
        SheetData(
            name=cols.SHEET_VIEW5,
            headers=cols.VIEW5_DEVICE_HEADERS,
            rows=[],
            numeric=set(cols.NUMERIC_HEADERS),
        ),
        SheetData(
            name=cols.SHEET_SCHEMA,
            headers=cols.PLACEMENT_HEADERS,
            rows=[placement(cols.ZONE_SCHEMA, 300)],
            numeric=set(cols.NUMERIC_HEADERS),
        ),
        SheetData(
            name=cols.SHEET_FOOTER,
            headers=cols.PLACEMENT_HEADERS,
            rows=[placement(cols.ZONE_FOOTER, 32)],
            numeric=set(cols.NUMERIC_HEADERS),
        ),
        SheetData(
            name=cols.SHEET_CONNECTIONS,
            headers=cols.CONNECTION_HEADERS,
            rows=[],
            numeric=set(cols.NUMERIC_HEADERS),
        ),
        SheetData(
            name=cols.SHEET_TEXTS,
            headers=cols.TEXT_HEADERS,
            rows=[text_example],
            numeric=set(cols.TEXT_NUMERIC_HEADERS),
        ),
        SheetData(
            name=cols.SHEET_SHEETS,
            headers=cols.SHEET_HEADERS,
            rows=[sheet_example],
            numeric=set(cols.SHEET_NUMERIC_HEADERS),
        ),
    ]
    return write_workbook(path, sheets)
